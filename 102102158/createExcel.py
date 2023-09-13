from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def createExcel():
    danmu_list = []  # 弹幕文件遍历
    f = open("弹幕.txt", "r", encoding="utf-8")
    for line in f.readlines():
        line = line.strip("\n")
        danmu_list.append(line)

    # 统计、排序
    danmu_counter = Counter(danmu_list)
    sorted_danmu = sorted(danmu_counter.items(), key=lambda x: x[1], reverse=True)
    for i in range(20):
        print(sorted_danmu[i])
        # print(sorted_danmu[i][0])

    # 创建一个新的Excel工作簿
    workbook = Workbook()
    sheet = workbook.active

    # 写入表头
    headers = ['弹幕', '数量']
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # 写入数据
    data = []
    for i in range(20):
        danmu, count = sorted_danmu[i]
        data.append([danmu, count])

    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_data in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_data)

    # 调整列宽
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 20

    # 保存Excel文件
    workbook.save('弹幕.xlsx')

